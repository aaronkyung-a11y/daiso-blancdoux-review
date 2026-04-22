import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

products = json.load(open('report_data_v2.json','r',encoding='utf-8'))
brand = json.load(open('brand_summary_v2.json','r',encoding='utf-8'))
attrs_all = json.load(open('attrs_by_product.json','r',encoding='utf-8'))

# Attach attributes to each product + compute pcts
for p in products:
    p['attributes'] = attrs_all.get(p['pdNo'], {}).get('attributes', [])
    tot = p['skincare_total'] + p['makeup_total']
    p['makeup_pct'] = round(p['makeup_total']/tot*100, 1) if tot else 0
    p['skincare_pct'] = round(p['skincare_total']/tot*100, 1) if tot else 0

wb = Workbook()
FONT='Malgun Gothic'
HEADER_FILL=PatternFill('solid',start_color='2C2C2C')
HEADER_FONT=Font(name=FONT,bold=True,color='FFFFFF',size=11)
TITLE_FONT=Font(name=FONT,bold=True,size=14,color='2C2C2C')
SUBTITLE_FONT=Font(name=FONT,bold=True,size=12,color='6B5B4F')
SK_FILL=PatternFill('solid',start_color='DFE8EF')
MK_FILL=PatternFill('solid',start_color='F2DCE2')
ACCENT_FILL=PatternFill('solid',start_color='F5EFE7')
BORDER=Border(left=Side(style='thin',color='DDDDDD'),right=Side(style='thin',color='DDDDDD'),
              top=Side(style='thin',color='DDDDDD'),bottom=Side(style='thin',color='DDDDDD'))
CENTER=Alignment(horizontal='center',vertical='center',wrap_text=True)
LEFT=Alignment(horizontal='left',vertical='center',wrap_text=True)
RIGHT=Alignment(horizontal='right',vertical='center')

# ===== Sheet 1: 개요 =====
ws=wb.active; ws.title='개요'
ws['A1']='더랩 바이 블랑두 — 다이소몰 VOC 키워드 분석 (v2)'
ws['A1'].font=Font(name=FONT,bold=True,size=16,color='2C2C2C')
ws.merge_cells('A1:G1')
ws['A2']=f'v2 업데이트: 중복 키워드 병합 (촉촉/촉촉하다→촉촉 등) + 스킨케어/메이크업 카테고리 분할  ·  분석일: 2026-04-22'
ws['A2'].font=Font(name=FONT,size=10,color='888888',italic=True)
ws.merge_cells('A2:G2')

# Product summary with makeup% 
ws['A4']='제품별 요약 (메이크업 비중 기준 정렬 가능)'
ws['A4'].font=SUBTITLE_FONT
headers=['pdNo','제품명','평점','리뷰수','스킨케어 언급','메이크업 언급','메이크업%']
for i,h in enumerate(headers,1):
    c=ws.cell(row=5,column=i,value=h)
    c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=CENTER; c.border=BORDER
for idx,p in enumerate(products,start=6):
    row=[p['pdNo'],p['pdNm'],float(p['avgRating']),int(p['revwCnt']),
         p['skincare_total'],p['makeup_total'],None]
    for i,v in enumerate(row,1):
        c=ws.cell(row=idx,column=i,value=v)
        c.font=Font(name=FONT,size=10); c.border=BORDER
        c.alignment=LEFT if i==2 else CENTER if i in (1,3,4) else RIGHT
    # makeup% as formula
    ws.cell(row=idx,column=7,value=f'=F{idx}/(E{idx}+F{idx})').number_format='0.0%'
    ws.cell(row=idx,column=7).font=Font(name=FONT,size=10,bold=(p['makeup_pct']>30))
    # highlight makeup-tilted rows
    if p['makeup_pct'] > 30:
        for i in range(1,8):
            ws.cell(row=idx,column=i).fill=MK_FILL
    elif idx % 2 == 0:
        for i in range(1,8):
            ws.cell(row=idx,column=i).fill=ACCENT_FILL
    ws.cell(row=idx,column=5).number_format='#,##0'
    ws.cell(row=idx,column=6).number_format='#,##0'

total_row=6+len(products)
ws.cell(row=total_row,column=1,value='브랜드 전체').font=Font(name=FONT,bold=True)
ws.cell(row=total_row,column=3,value=f'=AVERAGE(C6:C{total_row-1})').number_format='0.00'
ws.cell(row=total_row,column=4,value=f'=SUM(D6:D{total_row-1})').number_format='#,##0'
ws.cell(row=total_row,column=5,value=f'=SUM(E6:E{total_row-1})').number_format='#,##0'
ws.cell(row=total_row,column=6,value=f'=SUM(F6:F{total_row-1})').number_format='#,##0'
ws.cell(row=total_row,column=7,value=f'=F{total_row}/(E{total_row}+F{total_row})').number_format='0.0%'
for col in range(1,8):
    ws.cell(row=total_row,column=col).font=Font(name=FONT,bold=True,size=10)
    ws.cell(row=total_row,column=col).fill=PatternFill('solid',start_color='D4C5B0')

ws.column_dimensions['A'].width=14
ws.column_dimensions['B'].width=44
for col in 'CDEFG':
    ws.column_dimensions[col].width=13

# Brand block
start=total_row+3
ws.cell(row=start,column=1,value='브랜드 전체 키워드 집계').font=SUBTITLE_FONT
block_header=['분류','키워드/패턴','언급수']
for i,h in enumerate(block_header,1):
    c=ws.cell(row=start+1,column=i,value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=CENTER

r=start+2
for label,items,fill in [
    ('[스킨케어 언어]', brand['top_skincare_brand'][:15], SK_FILL),
    ('[메이크업 언어]', brand['top_makeup_brand'][:12], MK_FILL),
    ('[유입/Discovery]', brand['top_discovery_brand'], ACCENT_FILL),
    ('[긍정 키워드]',   brand['top_positive_brand'][:15], None),
    ('[불만 패턴]',    brand['top_complaints_brand'], None),
]:
    for i,(t,n) in enumerate(items):
        ws.cell(row=r,column=1,value=label if i==0 else '').font=Font(name=FONT,bold=(i==0),size=10)
        ws.cell(row=r,column=2,value=t).font=Font(name=FONT,size=10)
        ws.cell(row=r,column=3,value=n)
        ws.cell(row=r,column=3).number_format='#,##0'
        if fill:
            for col in [1,2,3]: ws.cell(row=r,column=col).fill=fill
        r+=1
    r+=1

# ===== Sheet 2: 카테고리 분할 (new) =====
ws=wb.create_sheet('카테고리분할')
ws['A1']='스킨케어 vs 메이크업 카테고리 분할'
ws['A1'].font=TITLE_FONT; ws.merge_cells('A1:G1')
ws['A2']=f'브랜드 전체: 스킨케어 언급 {brand["skincare_total"]:,}건 ({brand["skincare_share"]}%)  vs  메이크업 언급 {brand["makeup_total"]:,}건 ({brand["makeup_share"]}%)'
ws['A2'].font=Font(name=FONT,size=10,italic=True,color='6B5B4F'); ws.merge_cells('A2:G2')

ws['A4']='제품'; ws['B4']='리뷰수'
ws['C4']='스킨케어 TOP 키워드'; ws['D4']='스킨케어 언급'
ws['E4']='메이크업 TOP 키워드'; ws['F4']='메이크업 언급'; ws['G4']='메이크업%'
for col in range(1,8):
    c=ws.cell(row=4,column=col)
    c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=CENTER; c.border=BORDER

r=5
for p in products:
    start_r = r
    sk = p['skincare_keywords'][:6]
    mk = p['makeup_keywords'][:6]
    nrows = max(len(sk),len(mk),1)
    # Product cells merged across rows
    ws.cell(row=r,column=1,value=p['pdNm']).font=Font(name=FONT,size=10,bold=True)
    ws.cell(row=r,column=2,value=int(p['revwCnt'])).number_format='#,##0'
    if nrows>1:
        ws.merge_cells(start_row=r,start_column=1,end_row=r+nrows-1,end_column=1)
        ws.merge_cells(start_row=r,start_column=2,end_row=r+nrows-1,end_column=2)
    for i in range(nrows):
        sk_kw = sk[i] if i<len(sk) else None
        mk_kw = mk[i] if i<len(mk) else None
        if sk_kw:
            ws.cell(row=r+i,column=3,value=sk_kw['term']).font=Font(name=FONT,size=10)
            ws.cell(row=r+i,column=4,value=sk_kw['count']).number_format='#,##0'
            ws.cell(row=r+i,column=3).fill=SK_FILL
            ws.cell(row=r+i,column=4).fill=SK_FILL
        if mk_kw:
            ws.cell(row=r+i,column=5,value=mk_kw['term']).font=Font(name=FONT,size=10)
            ws.cell(row=r+i,column=6,value=mk_kw['count']).number_format='#,##0'
            ws.cell(row=r+i,column=5).fill=MK_FILL
            ws.cell(row=r+i,column=6).fill=MK_FILL
    # Makeup pct in first row
    mkpct_val = p['makeup_total']/(p['skincare_total']+p['makeup_total']) if (p['skincare_total']+p['makeup_total'])>0 else 0
    ws.cell(row=start_r,column=7,value=mkpct_val).number_format='0.0%'
    ws.cell(row=start_r,column=7).font=Font(name=FONT,size=11,bold=True,color='A8433A' if p['makeup_pct']>30 else '2C2C2C')
    ws.cell(row=start_r,column=7).alignment=Alignment(horizontal='center',vertical='center')
    if nrows>1:
        ws.merge_cells(start_row=start_r,start_column=7,end_row=start_r+nrows-1,end_column=7)
    # borders
    for rr in range(start_r, start_r+nrows):
        for cc in range(1,8):
            ws.cell(row=rr,column=cc).border=BORDER
            if cc in (1,2): ws.cell(row=rr,column=cc).alignment=Alignment(vertical='center',wrap_text=True)
    r += nrows
    # Small separator
    r += 0

ws.column_dimensions['A'].width=44
ws.column_dimensions['B'].width=9
ws.column_dimensions['C'].width=20
ws.column_dimensions['D'].width=12
ws.column_dimensions['E'].width=20
ws.column_dimensions['F'].width=12
ws.column_dimensions['G'].width=12
ws.freeze_panes='A5'

# ===== Sheet per product =====
for p in products:
    name = p['pdNm'].replace('더랩 바이 블랑두 ','').replace('더랩바이블랑두 ','')[:28]
    name = name.replace('/','-')
    ws = wb.create_sheet(name)
    ws['A1']=p['pdNm']; ws['A1'].font=TITLE_FONT; ws.merge_cells('A1:F1')
    ws['A2']=f"pdNo {p['pdNo']}  ·  ★{p['avgRating']}  ·  리뷰 {int(p['revwCnt']):,}건  ·  메이크업 비중 {p['makeup_pct']}%"
    ws['A2'].font=Font(name=FONT,size=10,color='888888',italic=True); ws.merge_cells('A2:F2')
    row=4

    # Skincare/Makeup split
    ws.cell(row=row,column=1,value='🔬 스킨케어 vs 💄 메이크업 — 고객이 쓴 언어').font=SUBTITLE_FONT
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6); row+=1
    ws.cell(row=row,column=1,value='Skincare 키워드').font=HEADER_FONT
    ws.cell(row=row,column=1).fill=HEADER_FILL
    ws.cell(row=row,column=2,value='언급').font=HEADER_FONT
    ws.cell(row=row,column=2).fill=HEADER_FILL
    ws.cell(row=row,column=4,value='Makeup 키워드').font=HEADER_FONT
    ws.cell(row=row,column=4).fill=HEADER_FILL
    ws.cell(row=row,column=5,value='언급').font=HEADER_FONT
    ws.cell(row=row,column=5).fill=HEADER_FILL
    for col in [1,2,4,5]:
        ws.cell(row=row,column=col).alignment=CENTER; ws.cell(row=row,column=col).border=BORDER
    row+=1
    sk = p['skincare_keywords'][:12]; mk = p['makeup_keywords'][:12]
    for i in range(max(len(sk),len(mk))):
        if i<len(sk):
            ws.cell(row=row+i,column=1,value=sk[i]['term']).font=Font(name=FONT,size=10)
            ws.cell(row=row+i,column=2,value=sk[i]['count']).number_format='#,##0'
            ws.cell(row=row+i,column=1).fill=SK_FILL; ws.cell(row=row+i,column=2).fill=SK_FILL
            ws.cell(row=row+i,column=1).border=BORDER; ws.cell(row=row+i,column=2).border=BORDER
            ws.cell(row=row+i,column=2).alignment=RIGHT
        if i<len(mk):
            ws.cell(row=row+i,column=4,value=mk[i]['term']).font=Font(name=FONT,size=10)
            ws.cell(row=row+i,column=5,value=mk[i]['count']).number_format='#,##0'
            ws.cell(row=row+i,column=4).fill=MK_FILL; ws.cell(row=row+i,column=5).fill=MK_FILL
            ws.cell(row=row+i,column=4).border=BORDER; ws.cell(row=row+i,column=5).border=BORDER
            ws.cell(row=row+i,column=5).alignment=RIGHT
    row += max(len(sk),len(mk)) + 2

    # Structured attrs
    if p['attributes']:
        ws.cell(row=row,column=1,value='🧪 다이소 구조화 설문').font=SUBTITLE_FONT
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6); row+=1
        hdr=['질문','항목','선택','응답자','비율']
        for i,h in enumerate(hdr,1):
            c=ws.cell(row=row,column=i,value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=CENTER; c.border=BORDER
        row+=1
        for a in p['attributes']:
            for opt in sorted(a['options'],key=lambda x:-x['percent']):
                ws.cell(row=row,column=1,value=a['question']).font=Font(name=FONT,size=10)
                ws.cell(row=row,column=2,value=a['item']).font=Font(name=FONT,size=10)
                ws.cell(row=row,column=3,value=opt['label']).font=Font(name=FONT,size=10,bold=(opt['percent']>=50))
                ws.cell(row=row,column=4,value=opt['count']).number_format='#,##0'
                ws.cell(row=row,column=5,value=opt['percent']/100).number_format='0%'
                ws.cell(row=row,column=5).font=Font(name=FONT,size=10,bold=(opt['percent']>=50))
                for col in range(1,6):
                    ws.cell(row=row,column=col).border=BORDER
                    ws.cell(row=row,column=col).alignment=LEFT if col in (1,2,3) else RIGHT
                    if row%2==0: ws.cell(row=row,column=col).fill=ACCENT_FILL
                row+=1
        row+=2

    # Positive
    ws.cell(row=row,column=1,value='😍 만족 키워드 (★5 과다 출현)').font=SUBTITLE_FONT
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6); row+=1
    for i,h in enumerate(['키워드','카테고리','총 언급','긍정‰','부정‰'],1):
        c=ws.cell(row=row,column=i,value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=CENTER; c.border=BORDER
    row+=1
    for k in p['positive_keywords'][:15]:
        ws.cell(row=row,column=1,value=k['term']).font=Font(name=FONT,size=10,bold=True)
        cat = k.get('category','other')
        cat_label = {'skincare':'🔬 스킨케어','makeup':'💄 메이크업','other':'일반'}[cat]
        ws.cell(row=row,column=2,value=cat_label).font=Font(name=FONT,size=10)
        ws.cell(row=row,column=3,value=k['count']).number_format='#,##0'
        ws.cell(row=row,column=4,value=k['pos_rate_per_1k'])
        ws.cell(row=row,column=5,value=k['neg_rate_per_1k'])
        for col in range(1,6):
            ws.cell(row=row,column=col).border=BORDER
            ws.cell(row=row,column=col).alignment=LEFT if col<=2 else RIGHT
        # Color-code by category
        if cat == 'skincare':
            ws.cell(row=row,column=2).fill=SK_FILL
        elif cat == 'makeup':
            ws.cell(row=row,column=2).fill=MK_FILL
        row+=1
    row+=2

    # Complaints
    ws.cell(row=row,column=1,value='😟 불만 패턴').font=SUBTITLE_FONT
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6); row+=1
    for i,h in enumerate(['패턴','건수','리뷰 대비 %'],1):
        c=ws.cell(row=row,column=i,value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=CENTER; c.border=BORDER
    row+=1
    for k in p['complaint_patterns']:
        ws.cell(row=row,column=1,value=k['label']).font=Font(name=FONT,size=10)
        ws.cell(row=row,column=2,value=k['count']).number_format='#,##0'
        ws.cell(row=row,column=3,value=f'=B{row}/{int(p["revwCnt"])}').number_format='0.0%'
        for col in range(1,4):
            ws.cell(row=row,column=col).border=BORDER
            ws.cell(row=row,column=col).alignment=LEFT if col==1 else RIGHT
        row+=1
    row+=2

    # Discovery
    ws.cell(row=row,column=1,value='🔍 유입·인지 경로').font=SUBTITLE_FONT
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6); row+=1
    for i,h in enumerate(['경로/동기','건수','리뷰 대비 %'],1):
        c=ws.cell(row=row,column=i,value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=CENTER; c.border=BORDER
    row+=1
    for k in p['discovery_mentions']:
        ws.cell(row=row,column=1,value=k['label']).font=Font(name=FONT,size=10)
        ws.cell(row=row,column=2,value=k['count']).number_format='#,##0'
        ws.cell(row=row,column=3,value=f'=B{row}/{int(p["revwCnt"])}').number_format='0.0%'
        for col in range(1,4):
            ws.cell(row=row,column=col).border=BORDER
            ws.cell(row=row,column=col).alignment=LEFT if col==1 else RIGHT
        row+=1
    row+=2

    # Low rating samples
    if p['low_rating_sample']:
        ws.cell(row=row,column=1,value='💬 저평점 리뷰 샘플 (★1~3)').font=SUBTITLE_FONT
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6); row+=1
        for s in p['low_rating_sample'][:5]:
            ws.cell(row=row,column=1,value=f"★{s['score']}").font=Font(name=FONT,size=10,bold=True,color='C94A4A')
            ws.cell(row=row,column=2,value=s['text']).font=Font(name=FONT,size=9)
            ws.cell(row=row,column=2).alignment=Alignment(wrap_text=True,vertical='top')
            ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=6)
            ws.row_dimensions[row].height=42
            row+=1

    ws.column_dimensions['A'].width=22
    ws.column_dimensions['B'].width=16
    ws.column_dimensions['C'].width=18
    ws.column_dimensions['D'].width=16
    ws.column_dimensions['E'].width=14
    ws.column_dimensions['F'].width=12
    ws.freeze_panes='A4'

out='/mnt/user-data/outputs/daiso_blancdoux_voc_analysis.xlsx'
import os; os.makedirs(os.path.dirname(out),exist_ok=True)
wb.save(out)
print(f"Saved: {out}")
